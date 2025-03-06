import os
import shutil
import time
import datetime
import pandas as pd
from geopy.geocoders import Nominatim
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from tkinter import filedialog, Tk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



# Manually input the client folder path (copy & paste the path here)
save_path = r"C:\Users\sebas\OneDrive - Atlas Technology Limited\Engineering\Clients\25_07_Habitat_4 Sunvale\Documents"

def log_message(message):
    with open("log.txt", "a", encoding="utf-8") as log_file:  # Force UTF-8 encoding
        timestamp = datetime.datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
        log_file.write(f"{timestamp} {message}\n")

# Ensure the path exists
if not os.path.exists(save_path):
    print(f"⚠ Error: The path '{save_path}' does not exist. Please check it.")
    exit()
else:
    print(f"✅ Using client folder: {save_path}")

def get_lat_long(address):
    geolocator = Nominatim(user_agent="hirds_automation")
    try:
        location = geolocator.geocode(address, timeout=10)
        if location:
            return location.latitude, location.longitude
        else:
            return None, None
    except Exception as e:
        print(f"Error: {e}")
        return None, None

# HIRDS Download Function
def open_hirds(lat, lon, save_path):
    print("Initializing Chrome WebDriver...")

    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        url = "https://hirds.niwa.co.nz/"
        driver.get(url)
        time.sleep(3)

        # Enter Latitude & Longitude
        driver.find_element(By.NAME, "latitude").clear()
        driver.find_element(By.NAME, "latitude").send_keys(str(lat))
        driver.find_element(By.NAME, "longitude").clear()
        driver.find_element(By.NAME, "longitude").send_keys(str(lon))

        # Select Depth-Duration and Generate Report
        driver.find_element(By.ID, "optionsRadios1").click()
        driver.find_element(By.XPATH, "//button[contains(@class, 'btn-primary')]").click()
        print("Generating HIRDS report...")
        time.sleep(5)

        # Click the Download Button
        driver.find_element(By.CLASS_NAME, "glyphicon-download-alt").click()
        print("Download started...")
        time.sleep(5)  # Wait for the download to complete
        driver.quit()

        # Find the latest downloaded file
        downloads_folder = os.path.expanduser("~\\Downloads")
        latest_file = max([os.path.join(downloads_folder, f) for f in os.listdir(downloads_folder)], key=os.path.getctime)

        # Move the file to the selected client folder
        hirds_destination = os.path.join(save_path, "HIRDS_Data.csv")
        shutil.move(latest_file, hirds_destination)
        print(f"✅ HIRDS data saved to: {hirds_destination}")

        return hirds_destination

    except Exception as e:
        print(f"ERROR: {e}")
        driver.quit()

# Copy Template Function
def copy_template(save_path):
    template_folder = r"C:\Users\sebas\OneDrive - Atlas Technology Limited\Engineering\Engineering Notes\1.0 Stormwater\Calculators\HEC Rainfall Assessment"
    latest_template = max([os.path.join(template_folder, f) for f in os.listdir(template_folder)], key=os.path.getctime)

    # Copy & Rename Template File
    template_destination = os.path.join(save_path, "HIRDS_Working.xlsx")
    shutil.copy(latest_template, template_destination)
    print(f"✅ Copied template to: {template_destination}")

    return template_destination

# **🔹 Transpose Required Data in "HIRDS Working Data"**
def transpose_hirds_data(sheet):
    print("🔄 Transposing HIRDS data...")

    # D7:J7 → M6:M12
    for i, col_idx in enumerate(range(4, 11)):  # Columns D to J (4 to 10)
        sheet[f"M{6 + i}"] = sheet[f"{get_column_letter(col_idx)}7"].value

    # D9:J9 → N6:N12
    for i, col_idx in enumerate(range(4, 11)):
        sheet[f"N{6 + i}"] = sheet[f"{get_column_letter(col_idx)}9"].value

    # D22:J22 → M21:M27
    for i, col_idx in enumerate(range(4, 11)):
        sheet[f"M{21 + i}"] = sheet[f"{get_column_letter(col_idx)}22"].value

    # D24:J24 → N21:N27
    for i, col_idx in enumerate(range(4, 11)):
        sheet[f"N{21 + i}"] = sheet[f"{get_column_letter(col_idx)}24"].value

    print("✅ Transpose complete!")


def process_hirds_data(hirds_file, template_file):
    print("🚀 Processing HIRDS Data...")

    if not os.path.exists(hirds_file):
        print(f"❌ ERROR: HIRDS file '{hirds_file}' not found.")
        return
    if not os.path.exists(template_file):
        print(f"❌ ERROR: Template file '{template_file}' not found.")
        return

    # **🔹 Load CSV & Skip Metadata**
    try:
        hirds_df = pd.read_csv(hirds_file, skiprows=11, delimiter=",")
    except pd.errors.ParserError:
        print("⚠ CSV Parsing Error! Trying with semicolon separator...")
        hirds_df = pd.read_csv(hirds_file, skiprows=11, delimiter=";")

    print(f"✅ HIRDS Data Loaded. Shape: {hirds_df.shape}")

    # **🔹 Extract Relevant Data**
    historical_data = hirds_df.iloc[:12, :9].values.tolist()  # A3 to I14 equivalent
    future_data = hirds_df.iloc[98:110, :9].values.tolist()  # A101 to I112 equivalent

    # **🔹 Load the Excel template**
    wb = load_workbook(template_file)
    working_sheet = wb["HIRDS Working Data"]

    # **🔹 Write Historical Data to "HIRDS Working Data" (B6 to J17)**
    for r_idx, row in enumerate(historical_data, start=6):
        for c_idx, value in enumerate(row, start=2):  # Start at column B (2)
            working_sheet.cell(row=r_idx, column=c_idx, value=value)

    # **🔹 Write Future Data (RCP6.0 2081-2100) to "HIRDS Working Data" (B21 to J32)**
    for r_idx, row in enumerate(future_data, start=21):
        for c_idx, value in enumerate(row, start=2):  # Start at column B (2)
            working_sheet.cell(row=r_idx, column=c_idx, value=value)

    # **🔹 Perform Transpose Operation**
    transpose_hirds_data(working_sheet)

    # **🔹 Add Site Address to "Site Information" (C5)**
    site_info_sheet = wb["Site Information"]
    site_info_sheet["C5"] = address  # Paste the address in C5
    print(f"✅ Address '{address}' added to Site Information (C5)")


    # **🔹 Save updated workbook**
    wb.save(template_file)
    print("✅ HIRDS data successfully transferred & transposed in the template!")


# Main Execution
if __name__ == "__main__":

    # Test Address
    #address = input("Please enter an address")
    address = "4 Sunvale Crescent, Whataupoko, Gisborne, New Zealand"
    lat, lon = get_lat_long(address)
    if lat and lon:
        print(f"Latitude: {lat}, Longitude: {lon}")  # Debug print
        log_message(f"SUCCESS: {address} → {lat}, {lon}")
    else:
        print("Failed to get coordinates.")
        log_message(f"ERROR: Failed to get coordinates for {address}")

    # Run HIRDS Download
    hirds_file = open_hirds(lat, lon, save_path)

    # Copy Template
    template_file = copy_template(save_path)

    # Process HIRDS Data & Update Template
    process_hirds_data(hirds_file, template_file)

    print("✅ All HIRDS processing complete!")
